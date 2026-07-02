import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import urllib.error
import urllib.request
import uuid
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr, Field

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
DB_PATH = Path(os.getenv("DATABASE_PATH", ROOT / "data" / "shop.sqlite3"))
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

FRONTEND = ROOT / "frontend"
UPLOAD_PRODUCTS = FRONTEND / "uploads" / "products"
UPLOAD_SITE = FRONTEND / "uploads" / "site"
UPLOAD_PRODUCTS.mkdir(parents=True, exist_ok=True)
UPLOAD_SITE.mkdir(parents=True, exist_ok=True)

OLLAMA_URL = os.getenv("OLLAMA_URL", "http://127.0.0.1:11434").rstrip("/")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.1:8b")
OLLAMA_TIMEOUT = int(os.getenv("OLLAMA_TIMEOUT", "180"))


class UserRegister(BaseModel):
    name: str = Field(min_length=2, max_length=80)
    email: EmailStr
    password: str = Field(min_length=6, max_length=120)
    phone: str = Field(default="", max_length=20)


class UserLogin(BaseModel):
    email: EmailStr
    password: str


class ProductIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    brand: str = Field(default="", max_length=100)
    category_id: int
    description: str = Field(default="", max_length=2000)
    characteristics: str = Field(default="", max_length=4000)
    composition_care: str = Field(default="", max_length=4000)
    price: float = Field(gt=0)
    compare_price: float | None = None
    image: str = Field(default="", max_length=500)
    images: str = Field(default="[]", max_length=1000)
    sizes: str = Field(default="[]", max_length=500)
    colors: str = Field(default="[]", max_length=500)
    tags: str = Field(default="[]", max_length=500)
    gender: str = Field(default="unisex", max_length=20)
    stock: int = Field(default=0, ge=0)
    active: bool = True
    featured: bool = False
    on_sale: bool = False
    rating: float = Field(default=4.5, ge=0, le=5)


class ProductQuickUpdate(BaseModel):
    active: bool | None = None
    on_sale: bool | None = None
    stock: int | None = Field(default=None, ge=0)
    price: float | None = Field(default=None, gt=0)


class CartItemIn(BaseModel):
    product_id: int
    quantity: int = Field(default=1, ge=1, le=99)
    size: str = Field(default="", max_length=50)
    color: str = Field(default="", max_length=50)


class CartItemUpdate(BaseModel):
    quantity: int | None = None
    size: str | None = None
    color: str | None = None


class CheckoutIn(BaseModel):
    address: str = Field(min_length=5, max_length=300)
    city: str = Field(min_length=2, max_length=100)
    state: str = Field(default="", max_length=100)
    zip_code: str = Field(default="", max_length=20)
    country: str = Field(default="ES", max_length=60)
    phone: str = Field(default="", max_length=20)
    notes: str = Field(default="", max_length=500)
    shipping_method: str = Field(default="standard", max_length=50)
    payment_method: str = Field(default="stripe", max_length=50)


class ChatIn(BaseModel):
    message: str = Field(min_length=1, max_length=2000)
    history: list[dict[str, str]] = Field(default_factory=list)
    context: str = Field(default="", max_length=12000)


app = FastAPI(title="MODE — Tienda de moda")
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/assets", StaticFiles(directory=FRONTEND), name="assets")


@contextmanager
def db() -> Any:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def row_to_dict(row: sqlite3.Row | None) -> dict[str, Any]:
    return dict(row) if row else {}


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9áéíóúñü]+", "-", text)
    return text.strip("-") or secrets.token_hex(4)


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    return f"{salt}:{hashlib.sha256((salt + password).encode()).hexdigest()}"


def verify_password(password: str, hashed: str) -> bool:
    try:
        salt, hsh = hashed.split(":", 1)
    except ValueError:
        return False
    return hsh == hashlib.sha256((salt + password).encode()).hexdigest()


def gen_token() -> str:
    return secrets.token_hex(32)


def get_user_by_token(token: str) -> dict | None:
    if not token:
        return None
    with db() as conn:
        row = conn.execute("SELECT * FROM users WHERE token = ?", (token,)).fetchone()
    return row_to_dict(row) if row else None


def require_admin(token: str) -> dict:
    user = get_user_by_token(token)
    if not user or user.get("role") != "admin":
        raise HTTPException(403, "Acceso denegado")
    return user


SEED_PRODUCTS = [
    {"name": "Chaqueta Aviator Premium", "brand": "Zara", "category": "Chaquetas", "description": "Chaqueta estilo aviador con acabado premium. Solapa amplia, cierre central y bolsillos laterales. Perfecta para looks casual elegantes.", "price": 89.95, "compare_price": 129.00, "image": "/assets/uploads/products/chaqueta-aviator-premium.jpg", "sizes": '["S","M","L","XL"]', "colors": '["Negro","Marrón","Verde"]', "tags": ["nuevo", "popular"], "gender": "hombre", "stock": 25},
    {"name": "Chaqueta de Cuero Clásica", "brand": "Mango", "category": "Chaquetas", "description": "Chaqueta de cuero genuino con corte clásico. Cremallera asimétrica y cuello con solapa. Un básico atemporal.", "price": 149.00, "compare_price": 199.00, "image": "/assets/uploads/products/chaqueta-cuero-clasica.jpg", "sizes": '["M","L","XL"]', "colors": '["Negro","Marrón"]', "tags": ["popular", "recomendado"], "gender": "hombre", "stock": 15},
    {"name": "Parka Invernal Algodón", "brand": "H&M", "category": "Chaquetas", "description": "Parka larga con capucha desmontable y forro térmico. Múltiples bolsillos y ajuste de cintura.", "price": 69.99, "compare_price": None, "image": "/assets/uploads/products/parka-invernal.jpg", "sizes": '["S","M","L","XL","XXL"]', "colors": '["Negro","Gris","Azul"]', "tags": ["oferta"], "gender": "unisex", "stock": 40, "on_sale": True},
    {"name": "Zapatillas Urbanas White", "brand": "Adidas", "category": "Zapatos", "description": "Zapatillas urbanas en piel blanca con suela ultraligera. Diseño minimalista y plantilla acolchada.", "price": 79.95, "compare_price": 99.95, "image": "/assets/uploads/products/zapatillas-urbanas-white.jpg", "sizes": '["39","40","41","42","43","44"]', "colors": '["Blanco","Negro"]', "tags": ["popular", "nuevo"], "gender": "unisex", "stock": 60},
    {"name": "Botas Trekking impermeables", "brand": "Columbia", "category": "Zapatos", "description": "Botas de trekking con membrana impermeable y suela Vibram. Ideales para montaña y senderismo.", "price": 129.00, "compare_price": 159.00, "image": "/assets/uploads/products/botas-trekking.jpg", "sizes": '["39","40","41","42","43"]', "colors": '["Gris","Verde","Negro"]', "tags": ["recomendado"], "gender": "unisex", "stock": 20},
    {"name": "Sneakers Retro Multicolor", "brand": "Nike", "category": "Zapatos", "description": "Sneakers de estilo retro con combinación de colores vibrantes. Suela gruesa y cierre de cordones.", "price": 99.00, "compare_price": None, "image": "/assets/uploads/products/sneakers-retro.jpg", "sizes": '["38","39","40","41","42","43","44"]', "colors": '["Blanco/Rojo","Negro/Dorado","Azul"]', "tags": ["nuevo"], "gender": "unisex", "stock": 35},
    {"name": "Pantalón Chino Slim Fit", "brand": "Zara", "category": "Pantalones", "description": "Pantalón chino de corte slim fit en algodón stretch. Cintura ajustable y costuras reforzadas.", "price": 39.95, "compare_price": 49.95, "image": "/assets/uploads/products/pantalon-chino.jpg", "sizes": '["30","32","34","36","38"]', "colors": '["Beige","Azul marino","Gris","Negro"]', "tags": ["popular"], "gender": "hombre", "stock": 50},
    {"name": "Jeans Skinny Mujer", "brand": "Levi's", "category": "Pantalones", "description": "Jeans skinny de tiro alto con elastano. Cinturilla ancha y costuras decorativas.", "price": 49.95, "compare_price": 59.95, "image": "/assets/uploads/products/jeans-skinny.jpg", "sizes": '["XS","S","M","L","XL"]', "colors": '["Azul claro","Azul oscuro","Negro"]', "tags": ["popular", "recomendado"], "gender": "mujer", "stock": 45},
    {"name": "Pantalón Cargo Holgado", "brand": "H&M", "category": "Pantalones", "description": "Pantalón cargo de corte holgado con múltiples bolsillos laterales. Cierre de cordón y goma en cintura.", "price": 34.99, "compare_price": None, "image": "/assets/uploads/products/pantalon-cargo.jpg", "sizes": '["S","M","L","XL","XXL"]', "colors": '["Verde oliva","Negro","Gris"]', "tags": ["nuevo", "oferta"], "gender": "hombre", "stock": 30, "on_sale": True},
    {"name": "Camiseta Algodón Premium", "brand": "H&M", "category": "Camisetas", "description": "Camiseta de algodón orgánico de 180 g/m². Corte regular y costuras planas. Imprescindible en cualquier armario.", "price": 19.95, "compare_price": 24.95, "image": "/assets/uploads/products/camiseta-algodon.jpg", "sizes": '["XS","S","M","L","XL","XXL"]', "colors": '["Blanco","Negro","Gris","Azul","Rojo"]', "tags": ["popular", "recomendado"], "gender": "unisex", "stock": 100},
    {"name": "Camiseta Estampada Vintage", "brand": "Mango", "category": "Camisetas", "description": "Camiseta con estampado vintage y acabado lavado. Mangas cortas y cuello redondo.", "price": 24.99, "compare_price": 29.99, "image": "/assets/uploads/products/camiseta-vintage.jpg", "sizes": '["S","M","L","XL"]', "colors": '["Negro","Blanco","Verde"]', "tags": ["nuevo"], "gender": "unisex", "stock": 55},
    {"name": "Polo Clásico Lisa", "brand": "Nike", "category": "Camisetas", "description": "Polo de manga corta con cuello tipo pico y dos botones. Tejido piqué de algodón.", "price": 34.95, "compare_price": 44.95, "image": "/assets/uploads/products/polo-clasico.jpg", "sizes": '["S","M","L","XL","XXL"]', "colors": '["Blanco","Negro","Azul marino","Rojo"]', "tags": ["popular"], "gender": "hombre", "stock": 70},
    {"name": "Mochila Urbana 25L", "brand": "Nike", "category": "Accesorios", "description": "Mochila urbana con compartimento para portátil de 15\". Bolsillo acolchado y correas ajustables.", "price": 44.95, "compare_price": 54.95, "image": "/assets/uploads/products/mochila-urbana.jpg", "sizes": '["Única"]', "colors": '["Negro","Gris","Azul"]', "tags": ["popular", "recomendado"], "gender": "unisex", "stock": 40},
    {"name": "Gorra Trucker Premium", "brand": "Adidas", "category": "Accesorios", "description": "Gorra estilo trucker con malla transpirable y visera curva. Cierre ajustable de clip.", "price": 24.99, "compare_price": 29.99, "image": "/assets/uploads/products/gorra-trucker.jpg", "sizes": '["Única"]', "colors": '["Negro","Blanco","Rojo","Azul"]', "tags": ["nuevo", "popular"], "gender": "unisex", "stock": 80},
    {"name": "Bufanda de Lana Merino", "brand": "Mango", "category": "Accesorios", "description": "Bufanda de lana merino 100% de 180 cm. Tejido suave y ribeteado. Ideal para el invierno.", "price": 29.95, "compare_price": 39.95, "image": "/assets/uploads/products/bufanda-lana.jpg", "sizes": '["Única"]', "colors": '["Gris","Negro","Burdeos","Verde"]', "tags": ["oferta"], "gender": "unisex", "stock": 35, "on_sale": True},
    {"name": "Cinturón Piel Vegana", "brand": "Zara", "category": "Accesorios", "description": "Cinturón de piel vegana de 3.5 cm de ancho. Hebilla metálica dorada/plateada. Ranuras cada 2.5 cm.", "price": 19.99, "compare_price": None, "image": "/assets/uploads/products/cinturon-piel.jpg", "sizes": '["80","85","90","95","100","105"]', "colors": '["Negro","Marrón"]', "tags": ["recomendado"], "gender": "hombre", "stock": 60},
    {"name": "Vestido Floral Midaxi", "brand": "Mango", "category": "Vestidos", "description": "Vestido midaxi con estampado floral. Escote en V y manga corta. Tejido ligero y fresco.", "price": 44.95, "compare_price": 54.95, "image": "/assets/uploads/products/vestido-floral.jpg", "sizes": '["XS","S","M","L","XL"]', "colors": '["Azul","Rojo","Negro"]', "tags": ["nuevo", "popular"], "gender": "mujer", "stock": 30},
    {"name": "Sudadera Oversize Capucha", "brand": "H&M", "category": "Sudaderas", "description": "Sudadera con capucha de corte oversize. Bolsillo canguro y costuras reforzadas. Tejido fleece.", "price": 39.95, "compare_price": 49.95, "image": "/assets/uploads/products/sudadera-oversize.jpg", "sizes": '["S","M","L","XL","XXL"]', "colors": '["Gris","Negro","Verde","Rosa"]', "tags": ["popular", "recomendado"], "gender": "unisex", "stock": 45},
    {"name": "Jersey Cuello Redondo", "brand": "Zara", "category": "Sudaderas", "description": "Jersey de cuello redondo en punto grueso. Ribetes elásticos y mangas largas. Cómodo y cálido.", "price": 34.99, "compare_price": None, "image": "/assets/uploads/products/jersey-cuello-redondo.jpg", "sizes": '["S","M","L","XL"]', "colors": '["Gris","Negro","Azul marino","Beige"]', "tags": ["oferta"], "gender": "unisex", "stock": 50, "on_sale": True},
    {"name": "Chaqueta Denim Clásica", "brand": "Levi's", "category": "Chaquetas", "description": "Chaqueta vaquera clásica con botones metálicos. Bolsillos con solapa y costuras en contraste.", "price": 69.95, "compare_price": 85.00, "image": "/assets/uploads/products/chaqueta-denim.jpg", "sizes": '["S","M","L","XL","XXL"]', "colors": '["Azul claro","Azul oscuro","Negro"]', "tags": ["popular"], "gender": "unisex", "stock": 25},
]

def init_db() -> None:
    with db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            phone TEXT DEFAULT '',
            role TEXT DEFAULT 'user',
            token TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            slug TEXT NOT NULL UNIQUE,
            image TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            brand TEXT DEFAULT '',
            category_id INTEGER NOT NULL,
            description TEXT DEFAULT '',
            price REAL NOT NULL,
            compare_price REAL,
            image TEXT DEFAULT '',
            images TEXT DEFAULT '[]',
            sizes TEXT DEFAULT '[]',
            colors TEXT DEFAULT '[]',
            tags TEXT DEFAULT '[]',
            gender TEXT DEFAULT 'unisex',
            stock INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            featured INTEGER DEFAULT 0,
            on_sale INTEGER DEFAULT 0,
            rating REAL DEFAULT 4.5,
            created_at TEXT NOT NULL,
            FOREIGN KEY(category_id) REFERENCES categories(id)
        );
        CREATE TABLE IF NOT EXISTS cart_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 1,
            size TEXT DEFAULT '',
            color TEXT DEFAULT '',
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(product_id) REFERENCES products(id)
        );
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            status TEXT DEFAULT 'pending',
            total REAL NOT NULL,
            subtotal REAL NOT NULL,
            tax REAL DEFAULT 0,
            shipping REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            address TEXT NOT NULL,
            city TEXT NOT NULL,
            state TEXT DEFAULT '',
            zip_code TEXT DEFAULT '',
            country TEXT DEFAULT 'ES',
            phone TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            shipping_method TEXT DEFAULT 'standard',
            payment_method TEXT DEFAULT 'stripe',
            payment_id TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER NOT NULL,
            product_id INTEGER NOT NULL,
            product_name TEXT NOT NULL,
            price REAL NOT NULL,
            quantity INTEGER NOT NULL,
            size TEXT DEFAULT '',
            color TEXT DEFAULT '',
            FOREIGN KEY(order_id) REFERENCES orders(id),
            FOREIGN KEY(product_id) REFERENCES products(id)
        );
        """)

        def ensure_column(table: str, column_def: str) -> None:
            column_name = column_def.split()[0]
            existing_cols = {r[1] for r in conn.execute(f"PRAGMA table_info({table})")}
            if column_name not in existing_cols:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column_def}")

        ensure_column("products", "characteristics TEXT DEFAULT ''")
        ensure_column("products", "composition_care TEXT DEFAULT ''")

        if not conn.execute("SELECT id FROM users WHERE email='admin@mode.com'").fetchone():
            conn.execute(
                "INSERT INTO users (name,email,password,role,token,created_at) VALUES (?,?,?,?,?,?)",
                ("Admin", "admin@mode.com", hash_password("admin123"), "admin", gen_token(), datetime.utcnow().isoformat()),
            )

        CATEGORY_IMAGES = {
            "Chaquetas": "/assets/uploads/site/chaquetas.jpg",
            "Zapatos": "/assets/uploads/site/zapatos.jpg",
            "Pantalones": "/assets/uploads/site/pantalones.jpg",
            "Camisetas": "/assets/uploads/site/camisetas.jpg",
            "Accesorios": "/assets/uploads/site/accesorios.jpg",
            "Vestidos": "/assets/uploads/site/vestidos.jpg",
            "Sudaderas": "/assets/uploads/site/sudaderas.jpg",
        }
        for name, img in CATEGORY_IMAGES.items():
            conn.execute("INSERT OR IGNORE INTO categories (name, slug, image) VALUES (?, ?, ?)", (name, slugify(name), img))
        cats = {r["name"]: r["id"] for r in conn.execute("SELECT id,name FROM categories")}

        existing = {r["name"] for r in conn.execute("SELECT name FROM products")}
        for p in SEED_PRODUCTS:
            if p["name"] not in existing:
                cat_id = cats.get(p.get("category", ""), 1)
                conn.execute(
                    """INSERT INTO products
                    (name, brand, category_id, description, price, compare_price, image,
                     sizes, colors, tags, gender, stock, active, featured, on_sale, rating, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,1,?,?,?,?)""",
                    (p["name"], p["brand"], cat_id, p["description"], p["price"],
                     p.get("compare_price"), p["image"], p["sizes"], p["colors"],
                     json.dumps(p.get("tags", [])), p["gender"], p["stock"],
                     1 if p.get("featured") else 0,
                     1 if p.get("on_sale") else 0,
                     p.get("rating", 4.5),
                     datetime.utcnow().isoformat())
                )


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/")
def index() -> FileResponse:
    return FileResponse(FRONTEND / "index.html")


@app.get("/sitemap.xml")
def sitemap() -> FileResponse:
    return FileResponse(FRONTEND / "sitemap.xml")


@app.get("/robots.txt")
def robots() -> FileResponse:
    return FileResponse(FRONTEND / "robots.txt")


@app.post("/api/auth/register")
def register(payload: UserRegister) -> dict:
    with db() as conn:
        if conn.execute("SELECT id FROM users WHERE email=?", (payload.email,)).fetchone():
            raise HTTPException(409, "El email ya está registrado")
        token = gen_token()
        conn.execute(
            "INSERT INTO users (name,email,password,phone,role,token,created_at) VALUES (?,?,?,?,?,?,?)",
            (payload.name, payload.email, hash_password(payload.password), payload.phone, "user", token, datetime.utcnow().isoformat()),
        )
        user = conn.execute("SELECT id,name,email,role FROM users WHERE token=?", (token,)).fetchone()
    return {**row_to_dict(user), "token": token}


@app.post("/api/auth/login")
def login(payload: UserLogin) -> dict:
    with db() as conn:
        user = conn.execute("SELECT * FROM users WHERE email=?", (payload.email,)).fetchone()
        if not user or not verify_password(payload.password, user["password"]):
            raise HTTPException(401, "Email o contraseña incorrectos")
        token = gen_token()
        conn.execute("UPDATE users SET token=? WHERE id=?", (token, user["id"]))
    return {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"], "token": token}


@app.get("/api/auth/me")
def me(token: str = Query("")) -> dict:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "No autenticado")
    return {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"]}


@app.get("/api/categories")
def categories() -> list[dict]:
    with db() as conn:
        rows = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/products")
def products(
    q: str = "", category: str = "", brand: str = "", min_price: float = 0,
    max_price: float = 9999, size: str = "", color: str = "", gender: str = "",
    tag: str = "", in_stock: bool = False, on_sale: bool = False,
    sort: str = "newest", page: int = Query(1, ge=1), per_page: int = Query(20, ge=1, le=100),
) -> dict:
    where = ["p.active=1"]
    params: list[Any] = []
    if q:
        where.append("(p.name LIKE ? OR p.brand LIKE ? OR p.description LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if category:
        where.append("c.slug=?")
        params.append(category)
    if brand:
        where.append("p.brand=?")
        params.append(brand)
    if min_price:
        where.append("p.price>=?")
        params.append(min_price)
    if max_price < 9999:
        where.append("p.price<=?")
        params.append(max_price)
    if size:
        where.append("p.sizes LIKE ?")
        params.append(f"%{size}%")
    if color:
        where.append("p.colors LIKE ?")
        params.append(f"%{color}%")
    if gender:
        where.append("p.gender IN (?, 'unisex')")
        params.append(gender)
    if tag:
        where.append("p.tags LIKE ?")
        params.append(f"%{tag}%")
    if in_stock:
        where.append("p.stock>0")
    if on_sale:
        where.append("p.on_sale=1")
    order = {
        "price_asc": "p.price ASC",
        "price_desc": "p.price DESC",
        "rating": "p.rating DESC",
        "oldest": "p.id ASC",
        "popular": "p.rating DESC, p.id DESC",
    }.get(sort, "p.id DESC")
    w = " AND ".join(where)
    offset = (page - 1) * per_page
    with db() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM products p JOIN categories c ON c.id=p.category_id WHERE {w}", params).fetchone()[0]
        rows = conn.execute(
            f"""SELECT p.*, c.name category_name, c.slug category_slug
            FROM products p JOIN categories c ON c.id=p.category_id
            WHERE {w} ORDER BY {order} LIMIT ? OFFSET ?""",
            params + [per_page, offset],
        ).fetchall()
    return {"items": [row_to_dict(r) for r in rows], "total": total, "page": page, "per_page": per_page}


@app.get("/api/products/{product_id}")
def product_detail(product_id: int) -> dict:
    with db() as conn:
        row = conn.execute(
            "SELECT p.*, c.name category_name, c.slug category_slug FROM products p JOIN categories c ON c.id=p.category_id WHERE p.id=? AND p.active=1",
            (product_id,),
        ).fetchone()
    if not row:
        raise HTTPException(404, "Producto no encontrado")
    return row_to_dict(row)


@app.post("/api/products")
def create_product(payload: ProductIn, token: str = Query("")) -> dict:
    require_admin(token)
    with db() as conn:
        conn.execute(
            """INSERT INTO products
            (name,brand,category_id,description,characteristics,composition_care,price,compare_price,image,images,sizes,colors,tags,gender,stock,active,featured,on_sale,rating,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (payload.name, payload.brand, payload.category_id, payload.description, payload.characteristics, payload.composition_care, payload.price, payload.compare_price,
             payload.image, payload.images, payload.sizes, payload.colors, payload.tags, payload.gender, payload.stock,
             int(payload.active), int(payload.featured), int(payload.on_sale), payload.rating, datetime.utcnow().isoformat()),
        )
        row = conn.execute("SELECT * FROM products ORDER BY id DESC LIMIT 1").fetchone()
    return row_to_dict(row)


@app.patch("/api/products/{product_id}")
def update_product(product_id: int, payload: ProductIn, token: str = Query("")) -> dict:
    require_admin(token)
    with db() as conn:
        conn.execute(
            """UPDATE products SET name=?,brand=?,category_id=?,description=?,characteristics=?,composition_care=?,price=?,compare_price=?,image=?,images=?,
            sizes=?,colors=?,tags=?,gender=?,stock=?,active=?,featured=?,on_sale=?,rating=? WHERE id=?""",
            (payload.name, payload.brand, payload.category_id, payload.description, payload.characteristics, payload.composition_care, payload.price, payload.compare_price, payload.image,
             payload.images, payload.sizes, payload.colors, payload.tags, payload.gender, payload.stock, int(payload.active),
             int(payload.featured), int(payload.on_sale), payload.rating, product_id),
        )
        row = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    return row_to_dict(row)


@app.delete("/api/products/{product_id}")
def delete_product(product_id: int, token: str = Query("")) -> dict:
    require_admin(token)
    with db() as conn:
        try:
            conn.execute("DELETE FROM products WHERE id=?", (product_id,))
            return {"ok": True, "deleted": True}
        except sqlite3.IntegrityError:
            conn.execute("UPDATE products SET active=0 WHERE id=?", (product_id,))
            return {"ok": True, "deactivated": True}


@app.get("/api/filters")
def filters() -> dict:
    with db() as conn:
        brands = [r["brand"] for r in conn.execute("SELECT DISTINCT brand FROM products WHERE active=1 AND brand!='' ORDER BY brand")]
        cats = [row_to_dict(r) for r in conn.execute("SELECT * FROM categories ORDER BY name")]
        sizes, colors = set(), set()
        for r in conn.execute("SELECT sizes,colors FROM products WHERE active=1"):
            for s in safe_json_list(r["sizes"]):
                sizes.add(s)
            for c in safe_json_list(r["colors"]):
                colors.add(c)
    return {"brands": brands, "categories": cats, "sizes": sorted(sizes), "colors": sorted(colors)}


def safe_json_list(value: str) -> list[str]:
    try:
        parsed = json.loads(value or "[]")
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []


def cart_rows(conn: sqlite3.Connection, user_id: int) -> list[dict]:
    rows = conn.execute(
        """SELECT ci.*, p.name, p.price, p.compare_price, p.image, p.stock, p.sizes, p.colors
        FROM cart_items ci JOIN products p ON p.id=ci.product_id WHERE ci.user_id=?""",
        (user_id,),
    ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/cart")
def get_cart(token: str = Query("")) -> list[dict]:
    user = get_user_by_token(token)
    if not user:
        return []
    with db() as conn:
        return cart_rows(conn, user["id"])


@app.post("/api/cart")
def add_to_cart(payload: CartItemIn, token: str = Query("")) -> list[dict]:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "Inicia sesión para añadir productos")
    with db() as conn:
        existing = conn.execute(
            "SELECT id FROM cart_items WHERE user_id=? AND product_id=? AND size=? AND color=?",
            (user["id"], payload.product_id, payload.size, payload.color),
        ).fetchone()
        if existing:
            conn.execute("UPDATE cart_items SET quantity=quantity+? WHERE id=?", (payload.quantity, existing["id"]))
        else:
            conn.execute("INSERT INTO cart_items (user_id,product_id,quantity,size,color) VALUES (?,?,?,?,?)", (user["id"], payload.product_id, payload.quantity, payload.size, payload.color))
        return cart_rows(conn, user["id"])


@app.patch("/api/cart/{item_id}")
def update_cart_item(item_id: int, payload: CartItemUpdate, token: str = Query("")) -> list[dict]:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "No autenticado")
    updates, values = [], []
    if payload.quantity is not None:
        updates.append("quantity=?")
        values.append(payload.quantity)
    if payload.size is not None:
        updates.append("size=?")
        values.append(payload.size)
    if payload.color is not None:
        updates.append("color=?")
        values.append(payload.color)
    with db() as conn:
        if updates:
            conn.execute(f"UPDATE cart_items SET {', '.join(updates)} WHERE id=? AND user_id=?", values + [item_id, user["id"]])
        return cart_rows(conn, user["id"])


@app.delete("/api/cart/{item_id}")
def remove_cart_item(item_id: int, token: str = Query("")) -> list[dict]:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "No autenticado")
    with db() as conn:
        conn.execute("DELETE FROM cart_items WHERE id=? AND user_id=?", (item_id, user["id"]))
        return cart_rows(conn, user["id"])


@app.post("/api/checkout")
def checkout(payload: CheckoutIn, token: str = Query("")) -> dict:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "Inicia sesión para comprar")
    with db() as conn:
        items = cart_rows(conn, user["id"])
        if not items:
            raise HTTPException(400, "El carrito está vacío")
        subtotal = sum(float(i["price"]) * int(i["quantity"]) for i in items)
        shipping = 0 if subtotal >= 50 else 4.99
        tax = round(subtotal * 0.21, 2)
        total = round(subtotal + shipping + tax, 2)
        now = datetime.utcnow().isoformat()
        conn.execute(
            """INSERT INTO orders (user_id,status,total,subtotal,tax,shipping,discount,address,city,state,zip_code,country,phone,notes,shipping_method,payment_method,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (user["id"], "pending", total, subtotal, tax, shipping, 0, payload.address, payload.city, payload.state, payload.zip_code, payload.country, payload.phone, payload.notes, payload.shipping_method, payload.payment_method, now, now),
        )
        order_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        for item in items:
            conn.execute("INSERT INTO order_items (order_id,product_id,product_name,price,quantity,size,color) VALUES (?,?,?,?,?,?,?)", (order_id, item["product_id"], item["name"], item["price"], item["quantity"], item["size"], item["color"]))
            conn.execute("UPDATE products SET stock=MAX(stock-?, 0) WHERE id=?", (item["quantity"], item["product_id"]))
        conn.execute("DELETE FROM cart_items WHERE user_id=?", (user["id"],))
    return {"order_id": order_id, "status": "pending", "total": total}


@app.get("/api/orders")
def orders_list(token: str = Query("")) -> list[dict]:
    user = get_user_by_token(token)
    if not user:
        raise HTTPException(401, "No autenticado")
    with db() as conn:
        rows = conn.execute("SELECT * FROM orders WHERE user_id=? ORDER BY created_at DESC", (user["id"],)).fetchall()
        result = []
        for row in rows:
            order = row_to_dict(row)
            order["items"] = [row_to_dict(i) for i in conn.execute("SELECT * FROM order_items WHERE order_id=?", (order["id"],))]
            result.append(order)
    return result


@app.get("/api/admin/products")
def admin_products(token: str = Query("")) -> list[dict]:
    require_admin(token)
    with db() as conn:
        rows = conn.execute("SELECT p.*, c.name category_name, c.slug category_slug FROM products p JOIN categories c ON c.id=p.category_id ORDER BY p.id DESC").fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/admin/product/{product_id}")
def admin_product_detail(product_id: int, token: str = Query("")) -> dict:
    require_admin(token)
    with db() as conn:
        row = conn.execute("SELECT p.*, c.name category_name, c.slug category_slug FROM products p JOIN categories c ON c.id=p.category_id WHERE p.id=?", (product_id,)).fetchone()
    if not row:
        raise HTTPException(404, "Producto no encontrado")
    return row_to_dict(row)


@app.patch("/api/admin/products/{product_id}/quick")
def quick_update_product(product_id: int, payload: ProductQuickUpdate, token: str = Query("")) -> dict:
    require_admin(token)
    updates, values = [], []
    if payload.active is not None:
        updates.append("active=?")
        values.append(int(payload.active))
    if payload.on_sale is not None:
        updates.append("on_sale=?")
        values.append(int(payload.on_sale))
    if payload.stock is not None:
        updates.append("stock=?")
        values.append(payload.stock)
    if payload.price is not None:
        updates.append("price=?")
        values.append(payload.price)
    if not updates:
        raise HTTPException(400, "No hay cambios para aplicar")
    with db() as conn:
        conn.execute(f"UPDATE products SET {', '.join(updates)} WHERE id=?", values + [product_id])
        row = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
    return row_to_dict(row)


@app.get("/api/admin/orders")
def admin_orders(token: str = Query(""), status: str = "") -> list[dict]:
    require_admin(token)
    where, params = ("WHERE status=?", [status]) if status else ("", [])
    with db() as conn:
        rows = conn.execute(f"SELECT * FROM orders {where} ORDER BY created_at DESC", params).fetchall()
        result = []
        for row in rows:
            order = row_to_dict(row)
            order["items"] = [row_to_dict(i) for i in conn.execute("SELECT * FROM order_items WHERE order_id=?", (order["id"],))]
            order["customer"] = row_to_dict(conn.execute("SELECT name,email FROM users WHERE id=?", (order["user_id"],)).fetchone())
            result.append(order)
    return result


@app.patch("/api/admin/orders/{order_id}/status")
def update_order_status(order_id: int, status: str = Query(""), token: str = Query("")) -> dict:
    require_admin(token)
    if status not in ("pending", "paid", "shipped", "delivered", "cancelled"):
        raise HTTPException(400, "Estado inválido")
    with db() as conn:
        conn.execute("UPDATE orders SET status=?, updated_at=? WHERE id=?", (status, datetime.utcnow().isoformat(), order_id))
        row = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    return row_to_dict(row)


@app.get("/api/admin/dashboard")
def admin_dashboard(token: str = Query("")) -> dict:
    require_admin(token)
    with db() as conn:
        revenue = conn.execute("SELECT COALESCE(SUM(total),0) FROM orders WHERE status IN ('paid','shipped','delivered')").fetchone()[0]
        revenue_by_status = {r["status"]: r["total"] for r in conn.execute("SELECT status, COALESCE(SUM(total),0) total FROM orders GROUP BY status")}
        orders_by_status = {r["status"]: r["cnt"] for r in conn.execute("SELECT status, COUNT(*) cnt FROM orders GROUP BY status")}
        cart_value = conn.execute("SELECT COALESCE(SUM(ci.quantity*p.price),0) FROM cart_items ci JOIN products p ON p.id=ci.product_id").fetchone()[0]
        data = {
            "total_revenue": revenue,
            "revenue_by_status": revenue_by_status,
            "orders_count_by_status": orders_by_status,
            "total_orders": conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0],
            "total_customers": conn.execute("SELECT COUNT(*) FROM users WHERE role='user'").fetchone()[0],
            "active_products": conn.execute("SELECT COUNT(*) FROM products WHERE active=1").fetchone()[0],
            "inactive_products": conn.execute("SELECT COUNT(*) FROM products WHERE active=0").fetchone()[0],
            "low_stock_products": [row_to_dict(r) for r in conn.execute("SELECT id,name,stock FROM products WHERE stock<=5 ORDER BY stock ASC LIMIT 10")],
            "cart_value": cart_value,
            "cart_item_count": conn.execute("SELECT COUNT(*) FROM cart_items").fetchone()[0],
            "customers_with_carts": conn.execute("SELECT COUNT(DISTINCT user_id) FROM cart_items").fetchone()[0],
            "recent_orders": [row_to_dict(r) for r in conn.execute("SELECT o.*, u.name customer_name, u.email customer_email FROM orders o JOIN users u ON u.id=o.user_id ORDER BY o.created_at DESC LIMIT 5")],
            "top_sold_products": [row_to_dict(r) for r in conn.execute("SELECT product_id, product_name, SUM(quantity) total_qty, SUM(quantity*price) total_rev FROM order_items GROUP BY product_id ORDER BY total_qty DESC LIMIT 5")],
            "most_carted_products": [row_to_dict(r) for r in conn.execute("SELECT p.id product_id, p.name, SUM(ci.quantity) total_qty, SUM(ci.quantity*p.price) cart_value FROM cart_items ci JOIN products p ON p.id=ci.product_id GROUP BY p.id ORDER BY cart_value DESC LIMIT 5")],
        }
    return data


@app.get("/api/admin/customers")
def admin_customers(token: str = Query("")) -> list[dict]:
    require_admin(token)
    with db() as conn:
        rows = conn.execute(
            """SELECT u.id,u.name,u.email,u.phone,u.created_at, COUNT(DISTINCT o.id) order_count,
            COALESCE(SUM(DISTINCT o.total),0) total_spent, COUNT(DISTINCT ci.id) cart_items,
            COALESCE(SUM(ci.quantity*p.price),0) cart_value
            FROM users u LEFT JOIN orders o ON o.user_id=u.id
            LEFT JOIN cart_items ci ON ci.user_id=u.id LEFT JOIN products p ON p.id=ci.product_id
            WHERE u.role='user' GROUP BY u.id ORDER BY total_spent DESC"""
        ).fetchall()
    return [row_to_dict(r) for r in rows]


@app.get("/api/admin/carts")
def admin_carts(token: str = Query("")) -> list[dict]:
    require_admin(token)
    with db() as conn:
        users = conn.execute(
            """SELECT u.id,u.name,u.email,COUNT(ci.id) item_count,COALESCE(SUM(ci.quantity*p.price),0) cart_value
            FROM users u JOIN cart_items ci ON ci.user_id=u.id JOIN products p ON p.id=ci.product_id
            GROUP BY u.id ORDER BY cart_value DESC"""
        ).fetchall()
        result = []
        for user in users:
            items = [row_to_dict(i) for i in conn.execute("SELECT ci.*, p.name,p.price,p.image,p.stock FROM cart_items ci JOIN products p ON p.id=ci.product_id WHERE ci.user_id=?", (user["id"],))]
            result.append({**row_to_dict(user), "items": items})
    return result


EXT_MAP = {"jpg": "jpg", "jpeg": "jpg", "png": "png", "webp": "webp", "avif": "avif", "gif": "gif"}


@app.post("/api/admin/uploads/product-image")
async def upload_product_image(token: str = Query(""), file: UploadFile = File(...)) -> dict:
    require_admin(token)
    if not (file.content_type or "").lower().startswith("image/"):
        raise HTTPException(400, "Solo se permiten imágenes")
    ext = EXT_MAP.get((file.filename or "").rsplit(".", 1)[-1].lower(), "")
    if not ext:
        raise HTTPException(400, "Formato no soportado")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(413, "La imagen supera 5 MB")
    name = f"{slugify(Path(file.filename or 'imagen').stem)}-{uuid.uuid4().hex[:10]}.{ext}"
    (UPLOAD_PRODUCTS / name).write_bytes(data)
    return {"url": f"/assets/uploads/products/{name}"}


XLSX_COLUMNS = ["name", "brand", "category_name", "description", "characteristics", "composition_care", "price", "compare_price", "image", "sizes", "colors", "tags", "gender", "stock", "active", "featured", "on_sale", "rating"]


@app.get("/api/admin/products/import-template")
def import_template(token: str = Query("")) -> StreamingResponse:
    require_admin(token)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(XLSX_COLUMNS)
    ws.append(["Camiseta Premium", "Nike", "Camisetas", "Descripción", "Marca: Nike\nCategoría: Camisetas\nGénero: unisex", "Algodón suave. Lavar a baja temperatura.", 29.95, 39.95, "", "S,M,L", "Negro,Blanco", "nuevo,popular", "unisex", 50, 1, 0, 1, 4.5])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"})


@app.get("/api/admin/products/export-excel")
def export_excel(token: str = Query("")) -> StreamingResponse:
    require_admin(token)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["id"] + XLSX_COLUMNS)
    with db() as conn:
        rows = conn.execute("SELECT p.*, c.name category_name FROM products p JOIN categories c ON c.id=p.category_id ORDER BY p.id").fetchall()
    for r in rows:
        d = row_to_dict(r)
        ws.append([d.get("id")] + [d.get(c, "") for c in XLSX_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=catalogo_productos.xlsx"})


def normalize_list(value: Any) -> str:
    if value is None:
        return "[]"
    raw = str(value).strip()
    if not raw:
        return "[]"
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return json.dumps(parsed, ensure_ascii=False)
    except Exception:
        pass
    return json.dumps([s.strip() for s in raw.split(",") if s.strip()], ensure_ascii=False)


@app.post("/api/admin/products/import-excel")
async def import_excel(token: str = Query(""), file: UploadFile = File(...)) -> dict:
    require_admin(token)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel vacío")
    headers = [str(h or "").strip().lower() for h in rows[0]]
    idx = {("category_name" if h == "category" else h): i for i, h in enumerate(headers)}
    if "name" not in idx or "price" not in idx:
        raise HTTPException(400, "Faltan columnas obligatorias: name, price")
    created = updated = skipped = 0
    errors: list[str] = []
    with db() as conn:
        cats = {r["name"].lower(): r["id"] for r in conn.execute("SELECT id,name FROM categories")}
        for n, row in enumerate(rows[1:], 2):
            try:
                get = lambda key, default="": row[idx[key]] if key in idx and idx[key] < len(row) and row[idx[key]] is not None else default
                name = str(get("name")).strip()
                if not name:
                    skipped += 1
                    continue
                brand = str(get("brand")).strip()
                cat = str(get("category_name", "General")).strip() or "General"
                if cat.lower() not in cats:
                    conn.execute("INSERT INTO categories (name,slug) VALUES (?,?)", (cat, slugify(cat)))
                    cats[cat.lower()] = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                payload = {
                    "brand": brand,
                    "category_id": cats[cat.lower()],
                    "description": str(get("description")),
                    "characteristics": str(get("characteristics", "")),
                    "composition_care": str(get("composition_care", "")),
                    "price": float(get("price")),
                    "compare_price": float(get("compare_price", 0)) or None,
                    "image": str(get("image", "")),
                    "sizes": normalize_list(get("sizes", "")),
                    "colors": normalize_list(get("colors", "")),
                    "tags": normalize_list(get("tags", "")),
                    "gender": str(get("gender", "unisex")),
                    "stock": int(float(get("stock", 0))),
                    "active": int(float(get("active", 1))),
                    "featured": int(float(get("featured", 0))),
                    "on_sale": int(float(get("on_sale", 0))),
                    "rating": float(get("rating", 4.5)),
                }
                existing = conn.execute("SELECT id FROM products WHERE name=? AND brand=?", (name, brand)).fetchone()
                if existing:
                    conn.execute(
                        """UPDATE products SET brand=?,category_id=?,description=?,characteristics=?,composition_care=?,price=?,compare_price=?,image=?,sizes=?,colors=?,tags=?,gender=?,stock=?,active=?,featured=?,on_sale=?,rating=? WHERE id=?""",
                        (payload["brand"], payload["category_id"], payload["description"], payload["characteristics"], payload["composition_care"], payload["price"], payload["compare_price"], payload["image"], payload["sizes"], payload["colors"], payload["tags"], payload["gender"], payload["stock"], payload["active"], payload["featured"], payload["on_sale"], payload["rating"], existing["id"]),
                    )
                    updated += 1
                else:
                    conn.execute("""INSERT INTO products (name,brand,category_id,description,characteristics,composition_care,price,compare_price,image,sizes,colors,tags,gender,stock,active,featured,on_sale,rating,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (name, payload["brand"], payload["category_id"], payload["description"], payload["characteristics"], payload["composition_care"], payload["price"], payload["compare_price"], payload["image"], payload["sizes"], payload["colors"], payload["tags"], payload["gender"], payload["stock"], payload["active"], payload["featured"], payload["on_sale"], payload["rating"], datetime.utcnow().isoformat()))
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Fila {n}: {exc}")
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


@app.post("/api/admin/products/ai-copy")
def admin_product_ai_copy(payload: ProductIn, token: str = Query("")) -> dict:
    require_admin(token)
    if not payload.name.strip() or not payload.brand.strip() or not payload.category_id:
        raise HTTPException(400, "Rellena nombre, marca y categoría antes de generar con IA")
    with db() as conn:
        cat = conn.execute("SELECT name FROM categories WHERE id=?", (payload.category_id,)).fetchone()
    category_name = cat["name"] if cat else "General"
    prompt = (
        "Genera copy para un producto de una tienda de moda en español. Devuelve EXACTAMENTE este formato, sin markdown ni texto extra:\n"
        "DESCRIPTION:\n<un párrafo breve y comercial>\n\n"
        "CHARACTERISTICS:\n<texto en líneas para el bloque Características>\n\n"
        "COMPOSITION_CARE:\n<texto breve de composición y cuidados>\n\n"
        "No copies el texto actual tal cual, mejora la redacción.\n\n"
        f"Nombre: {payload.name.strip()}\n"
        f"Marca: {payload.brand.strip()}\n"
        f"Categoría: {category_name}\n"
        f"Precio: {payload.price}\n"
        f"Género: {payload.gender}\n"
        f"Tags: {payload.tags}\n"
        f"Descripción actual: {payload.description.strip()}"
    )
    fallback = {
        "description": payload.description.strip() or f"{payload.name.strip()} de {payload.brand.strip()}, pensada para el día a día con estilo y comodidad.",
        "characteristics": f"Marca: {payload.brand.strip()}\nCategoría: {category_name}\nGénero: {payload.gender}\nPrecio: {payload.price:.2f} €",
        "composition_care": "Tejido de alta calidad. Consulta la etiqueta para instrucciones de lavado y cuidado.",
    }
    try:
        req = urllib.request.Request(
            f"{OLLAMA_URL}/api/generate",
            data=json.dumps({"model": OLLAMA_MODEL, "prompt": prompt, "stream": False, "options": {"num_predict": 500}}).encode(),
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=OLLAMA_TIMEOUT) as res:
            data = json.loads(res.read().decode())
        text = (data.get("response") or "").strip()
        parsed = {}
        for key, label in (("description", "DESCRIPTION"), ("characteristics", "CHARACTERISTICS"), ("composition_care", "COMPOSITION_CARE")):
            m = re.search(rf"{label}\s*:\s*(.*?)(?=\n[A-Z_ ]+\s*:\s*|\Z)", text, re.S | re.I)
            if m:
                parsed[key] = m.group(1).strip()
        if not parsed:
            parsed = json.loads(text)
        return {
            "description": str(parsed.get("description", fallback["description"])).strip(),
            "characteristics": str(parsed.get("characteristics", fallback["characteristics"])).strip(),
            "composition_care": str(parsed.get("composition_care", fallback["composition_care"])).strip(),
            "model": OLLAMA_MODEL,
        }
    except Exception:
        return {**fallback, "model": "fallback"}


@app.post("/api/ai/chat")
def ai_chat(payload: ChatIn, token: str = Query("")) -> dict:
    with db() as conn:
        products_ctx = "\n".join(
            f"- {r['name']} ({r['brand']}), {r['price']} EUR, stock {r['stock']}"
            for r in conn.execute("SELECT name,brand,price,stock FROM products WHERE active=1 ORDER BY rating DESC LIMIT 12")
        )
    history_lines = []
    for item in (payload.history or [])[-12:]:
        role = str(item.get("role", "")).strip().lower()
        content = str(item.get("content", "")).strip()
        if not content:
            continue
        if role not in ("user", "assistant"):
            role = "user"
        history_lines.append(f"{role}: {content}")
    prompt = "Eres asesor de moda de una tienda online. Responde en español, breve y útil. Recomienda solo productos del catálogo si encajan."
    if history_lines:
        prompt += "\n\nHistorial reciente de conversación:\n" + "\n".join(history_lines)
    if payload.context.strip():
        prompt += f"\n\nContexto interno para analizar la tienda:\n{payload.context.strip()}"
    prompt += f"\n\nCatálogo disponible:\n{products_ctx}\n\nPregunta del usuario: {payload.message}"
    fallback = "Ahora mismo puedo ayudarte con chaquetas, camisetas, zapatos, pantalones y accesorios. Dime talla, color, presupuesto y ocasión para recomendarte una opción."
    try:
        req = urllib.request.Request(
            f"{OLLAMA_URL}/api/generate",
            data=json.dumps({"model": OLLAMA_MODEL, "prompt": prompt, "stream": False, "options": {"num_predict": 500}}).encode(),
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=OLLAMA_TIMEOUT) as res:
            data = json.loads(res.read().decode())
        return {"response": data.get("response", fallback), "model": OLLAMA_MODEL}
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return {"response": fallback, "model": "fallback"}


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok", "shop": "MODE Clothing Store"}


@app.get("/{path:path}", response_model=None)
def spa(path: str):
    if path.startswith("api/") or path.startswith("assets/") or path in ("sitemap.xml", "robots.txt"):
        return JSONResponse({"detail": "Not Found"}, status_code=404)
    return FileResponse(FRONTEND / "index.html")
